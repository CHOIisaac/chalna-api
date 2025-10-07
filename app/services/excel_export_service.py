"""
📊 엑셀 내보내기 서비스

통계 데이터를 엑셀 파일로 내보내는 기능을 제공합니다.
"""

import io
from datetime import datetime
from typing import Dict, Any, List, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.core.database import get_db
from app.core.security import get_current_user
from app.api.stats import (
    get_monthly_stats,
    get_total_amounts,
    get_top_items,
    get_amount_distribution,
    get_relationship_breakdown,
    get_personal_details,
    get_events_stats
)


class ExcelExportService:
    """엑셀 내보내기 서비스 클래스"""

    def __init__(self):
        self.workbook = None
        self.styles = self._create_styles()

    def _create_styles(self) -> Dict[str, Any]:
        """엑셀 스타일 정의"""
        return {
            "header_font": Font(bold=True, color="FFFFFF"),
            "header_fill": PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
            "subheader_font": Font(bold=True, color="2F4F4F"),
            "subheader_fill": PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"),
            "center_alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            ),
            "currency_format": "#,##0"
        }

    async def export_all_stats(self, user_id: int, db) -> io.BytesIO:
        """모든 통계 데이터를 엑셀로 내보내기"""
        self.workbook = Workbook()

        # 기본 시트 제거
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])

        # 각 통계 데이터를 별도 시트로 생성
        await self._create_monthly_stats_sheet(user_id, db)
        await self._create_total_amounts_sheet(user_id, db)
        await self._create_top_items_sheet(user_id, db)
        await self._create_amount_distribution_sheet(user_id, db)
        await self._create_relationship_breakdown_sheet(user_id, db)
        await self._create_personal_details_sheet(user_id, db)
        await self._create_events_stats_sheet(user_id, db)

        # 메모리 버퍼에 저장
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)

        return output

    async def _create_monthly_stats_sheet(self, user_id: int, db):
        """월별 통계 시트 생성"""
        ws = self.workbook.create_sheet("월별 통계")

        # 데이터 조회
        monthly_data = await get_monthly_stats(user_id, db)
        data = monthly_data["data"]

        # 헤더 작성
        ws["A1"] = "월별 통계"
        ws["A1"].font = self.styles["header_font"]
        ws["A1"].fill = self.styles["header_fill"]
        ws.merge_cells("A1:F1")

        # 축의금 데이터
        row = 3
        ws[f"A{row}"] = "축의금 (나눔)"
        ws[f"A{row}"].font = self.styles["subheader_font"]
        ws[f"A{row}"].fill = self.styles["subheader_fill"]
        ws.merge_cells(f"A{row}:C{row}")

        row += 1
        ws[f"A{row}"] = "월"
        ws[f"B{row}"] = "금액"
        ws[f"C{row}"] = "비고"

        for item in data["wedding"]["given"]:
            row += 1
            ws[f"A{row}"] = item["month"]
            ws[f"B{row}"] = item["amount"]
            ws[f"C{row}"] = "나눔"

        # 축의금 받음 데이터
        row += 2
        ws[f"A{row}"] = "축의금 (받음)"
        ws[f"A{row}"].font = self.styles["subheader_font"]
        ws[f"A{row}"].fill = self.styles["subheader_fill"]
        ws.merge_cells(f"A{row}:C{row}")

        row += 1
        ws[f"A{row}"] = "월"
        ws[f"B{row}"] = "금액"
        ws[f"C{row}"] = "비고"

        for item in data["wedding"]["received"]:
            row += 1
            ws[f"A{row}"] = item["month"]
            ws[f"B{row}"] = item["amount"]
            ws[f"C{row}"] = "받음"

        # 조의금 데이터
        row += 2
        ws[f"A{row}"] = "조의금 (나눔)"
        ws[f"A{row}"].font = self.styles["subheader_font"]
        ws[f"A{row}"].fill = self.styles["subheader_fill"]
        ws.merge_cells(f"A{row}:C{row}")

        row += 1
        ws[f"A{row}"] = "월"
        ws[f"B{row}"] = "금액"
        ws[f"C{row}"] = "비고"

        for item in data["condolence"]["given"]:
            row += 1
            ws[f"A{row}"] = item["month"]
            ws[f"B{row}"] = item["amount"]
            ws[f"C{row}"] = "나눔"

        # 조의금 받음 데이터
        row += 2
        ws[f"A{row}"] = "조의금 (받음)"
        ws[f"A{row}"].font = self.styles["subheader_font"]
        ws[f"A{row}"].fill = self.styles["subheader_fill"]
        ws.merge_cells(f"A{row}:C{row}")

        row += 1
        ws[f"A{row}"] = "월"
        ws[f"B{row}"] = "금액"
        ws[f"C{row}"] = "비고"

        for item in data["condolence"]["received"]:
            row += 1
            ws[f"A{row}"] = item["month"]
            ws[f"B{row}"] = item["amount"]
            ws[f"C{row}"] = "받음"

        self._apply_styles_to_sheet(ws)

    async def _create_total_amounts_sheet(self, user_id: int, db):
        """총액 조회 시트 생성"""
        ws = self.workbook.create_sheet("총액 조회")

        # 데이터 조회
        total_data = await get_total_amounts(user_id, db)
        data = total_data["data"]

        # 헤더 작성
        ws["A1"] = "총액 조회"
        ws["A1"].font = self.styles["header_font"]
        ws["A1"].fill = self.styles["header_fill"]
        ws.merge_cells("A1:E1")

        # 테이블 헤더
        row = 3
        headers = ["구분", "축의금 총액", "축의금 건수", "조의금 총액", "조의금 건수"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles["subheader_font"]
            cell.fill = self.styles["subheader_fill"]

        # 나눔 데이터
        row += 1
        ws[f"A{row}"] = "나눔"
        ws[f"B{row}"] = data["given"]["wedding"]["total"]
        ws[f"C{row}"] = data["given"]["wedding"]["count"]
        ws[f"D{row}"] = data["given"]["condolence"]["total"]
        ws[f"E{row}"] = data["given"]["condolence"]["count"]

        # 받음 데이터
        row += 1
        ws[f"A{row}"] = "받음"
        ws[f"B{row}"] = data["received"]["wedding"]["total"]
        ws[f"C{row}"] = data["received"]["wedding"]["count"]
        ws[f"D{row}"] = data["received"]["condolence"]["total"]
        ws[f"E{row}"] = data["received"]["condolence"]["count"]

        self._apply_styles_to_sheet(ws)

    async def _create_top_items_sheet(self, user_id: int, db):
        """TOP 5 항목 시트 생성"""
        ws = self.workbook.create_sheet("TOP 5 항목")

        # 데이터 조회
        top_data = await get_top_items(5, user_id, db)
        data = top_data["data"]

        # 헤더 작성
        ws["A1"] = "TOP 5 항목"
        ws["A1"].font = self.styles["header_font"]
        ws["A1"].fill = self.styles["header_fill"]
        ws.merge_cells("A1:D1")

        # 나눔 TOP 5
        row = 3
        ws[f"A{row}"] = "나눔 TOP 5"
        ws[f"A{row}"].font = self.styles["subheader_font"]
        ws[f"A{row}"].fill = self.styles["subheader_fill"]
        ws.merge_cells(f"A{row}:D{row}")

        row += 1
        headers = ["순위", "이름", "금액", "구분"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles["subheader_font"]
            cell.fill = self.styles["subheader_fill"]

        for i, item in enumerate(data["given"], 1):
            row += 1
            ws[f"A{row}"] = i
            ws[f"B{row}"] = item["name"]
            ws[f"C{row}"] = item["amount"]
            ws[f"D{row}"] = item["type"]

        # 받음 TOP 5
        row += 2
        ws[f"A{row}"] = "받음 TOP 5"
        ws[f"A{row}"].font = self.styles["subheader_font"]
        ws[f"A{row}"].fill = self.styles["subheader_fill"]
        ws.merge_cells(f"A{row}:D{row}")

        row += 1
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles["subheader_font"]
            cell.fill = self.styles["subheader_fill"]

        for i, item in enumerate(data["received"], 1):
            row += 1
            ws[f"A{row}"] = i
            ws[f"B{row}"] = item["name"]
            ws[f"C{row}"] = item["amount"]
            ws[f"D{row}"] = item["type"]

        self._apply_styles_to_sheet(ws)

    async def _create_amount_distribution_sheet(self, user_id: int, db):
        """금액대별 분포 시트 생성"""
        ws = self.workbook.create_sheet("금액대별 분포")

        # 데이터 조회
        distribution_data = await get_amount_distribution(user_id, db)
        data = distribution_data["data"]

        # 헤더 작성
        ws["A1"] = "금액대별 분포"
        ws["A1"].font = self.styles["header_font"]
        ws["A1"].fill = self.styles["header_fill"]
        ws.merge_cells("A1:E1")

        # 나눔 분포
        row = 3
        ws[f"A{row}"] = "나눔 금액대별 분포"
        ws[f"A{row}"].font = self.styles["subheader_font"]
        ws[f"A{row}"].fill = self.styles["subheader_fill"]
        ws.merge_cells(f"A{row}:E{row}")

        row += 1
        headers = ["금액대", "건수", "비율(%)", "", ""]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles["subheader_font"]
            cell.fill = self.styles["subheader_fill"]

        for item in data["given"]:
            row += 1
            ws[f"A{row}"] = item["range"]
            ws[f"B{row}"] = item["count"]
            ws[f"C{row}"] = item["percentage"]

        # 받음 분포
        row += 2
        ws[f"A{row}"] = "받음 금액대별 분포"
        ws[f"A{row}"].font = self.styles["subheader_font"]
        ws[f"A{row}"].fill = self.styles["subheader_fill"]
        ws.merge_cells(f"A{row}:E{row}")

        row += 1
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles["subheader_font"]
            cell.fill = self.styles["subheader_fill"]

        for item in data["received"]:
            row += 1
            ws[f"A{row}"] = item["range"]
            ws[f"B{row}"] = item["count"]
            ws[f"C{row}"] = item["percentage"]

        self._apply_styles_to_sheet(ws)

    async def _create_relationship_breakdown_sheet(self, user_id: int, db):
        """관계별 분석 시트 생성"""
        ws = self.workbook.create_sheet("관계별 분석")

        # 데이터 조회
        relationship_data = await get_relationship_breakdown(user_id, db)
        data = relationship_data["data"]

        # 헤더 작성
        ws["A1"] = "관계별 분석"
        ws["A1"].font = self.styles["header_font"]
        ws["A1"].fill = self.styles["header_fill"]
        ws.merge_cells("A1:F1")

        # 나눔 관계별 분석
        row = 3
        ws[f"A{row}"] = "나눔 관계별 분석"
        ws[f"A{row}"].font = self.styles["subheader_font"]
        ws[f"A{row}"].fill = self.styles["subheader_fill"]
        ws.merge_cells(f"A{row}:F{row}")

        row += 1
        headers = ["관계", "건수", "총액", "평균금액", "", ""]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles["subheader_font"]
            cell.fill = self.styles["subheader_fill"]

        for item in data["given"]:
            row += 1
            ws[f"A{row}"] = item["relationship"]
            ws[f"B{row}"] = item["count"]
            ws[f"C{row}"] = item["totalAmount"]
            ws[f"D{row}"] = item["avgAmount"]

        # 받음 관계별 분석
        row += 2
        ws[f"A{row}"] = "받음 관계별 분석"
        ws[f"A{row}"].font = self.styles["subheader_font"]
        ws[f"A{row}"].fill = self.styles["subheader_fill"]
        ws.merge_cells(f"A{row}:F{row}")

        row += 1
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles["subheader_font"]
            cell.fill = self.styles["subheader_fill"]

        for item in data["received"]:
            row += 1
            ws[f"A{row}"] = item["relationship"]
            ws[f"B{row}"] = item["count"]
            ws[f"C{row}"] = item["totalAmount"]
            ws[f"D{row}"] = item["avgAmount"]

        self._apply_styles_to_sheet(ws)

    async def _create_personal_details_sheet(self, user_id: int, db):
        """개인별 상세 시트 생성"""
        ws = self.workbook.create_sheet("개인별 상세")

        # 데이터 조회
        personal_data = await get_personal_details(user_id, db)
        data = personal_data["data"]

        # 헤더 작성
        ws["A1"] = "개인별 상세"
        ws["A1"].font = self.styles["header_font"]
        ws["A1"].fill = self.styles["header_fill"]
        ws.merge_cells("A1:F1")

        # 나눔 개인별 상세
        row = 3
        ws[f"A{row}"] = "나눔 개인별 상세"
        ws[f"A{row}"].font = self.styles["subheader_font"]
        ws[f"A{row}"].fill = self.styles["subheader_fill"]
        ws.merge_cells(f"A{row}:F{row}")

        row += 1
        headers = ["이름", "총액", "건수", "평균금액", "관계", ""]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles["subheader_font"]
            cell.fill = self.styles["subheader_fill"]

        for item in data["given"]:
            row += 1
            ws[f"A{row}"] = item["name"]
            ws[f"B{row}"] = item["total"]
            ws[f"C{row}"] = item["count"]
            ws[f"D{row}"] = item["avg"]
            ws[f"E{row}"] = item["relationship"]

        # 받음 개인별 상세
        row += 2
        ws[f"A{row}"] = "받음 개인별 상세"
        ws[f"A{row}"].font = self.styles["subheader_font"]
        ws[f"A{row}"].fill = self.styles["subheader_fill"]
        ws.merge_cells(f"A{row}:F{row}")

        row += 1
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles["subheader_font"]
            cell.fill = self.styles["subheader_fill"]

        for item in data["received"]:
            row += 1
            ws[f"A{row}"] = item["name"]
            ws[f"B{row}"] = item["total"]
            ws[f"C{row}"] = item["count"]
            ws[f"D{row}"] = item["avg"]
            ws[f"E{row}"] = item["relationship"]

        self._apply_styles_to_sheet(ws)

    async def _create_events_stats_sheet(self, user_id: int, db):
        """이벤트별 기록 시트 생성"""
        ws = self.workbook.create_sheet("이벤트별 기록")

        # 데이터 조회
        events_data = await get_events_stats(user_id, db)
        data = events_data["data"]

        # 헤더 작성
        ws["A1"] = "이벤트별 기록"
        ws["A1"].font = self.styles["header_font"]
        ws["A1"].fill = self.styles["header_fill"]
        ws.merge_cells("A1:D1")

        # 테이블 헤더
        row = 3
        headers = ["이벤트 타입", "건수", "평균금액", "비고"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles["subheader_font"]
            cell.fill = self.styles["subheader_fill"]

        # 데이터 작성
        for item in data:
            row += 1
            ws[f"A{row}"] = item["type"]
            ws[f"B{row}"] = item["count"]
            ws[f"C{row}"] = item["avgAmount"]
            ws[f"D{row}"] = "통계"

        self._apply_styles_to_sheet(ws)

    def _apply_styles_to_sheet(self, ws):
        """시트에 스타일 적용"""
        # 열 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = None
            
            # 병합된 셀이 아닌 첫 번째 셀 찾기
            for cell in column:
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                    break
            
            if column_letter:
                for cell in column:
                    try:
                        if hasattr(cell, 'value') and cell.value is not None:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # 테두리 적용
        for row in ws.iter_rows():
            for cell in row:
                if hasattr(cell, 'value') and cell.value is not None:
                    cell.border = self.styles["border"]
                    if cell.row > 1:  # 헤더가 아닌 경우
                        cell.alignment = self.styles["center_alignment"]
    
    async def export_monthly_stats(self, user_id: int, db) -> io.BytesIO:
        """월별 통계만 엑셀로 내보내기"""
        self.workbook = Workbook()
        
        # 기본 시트 제거
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])
        
        await self._create_monthly_stats_sheet(user_id, db)
        
        # 메모리 버퍼에 저장
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        
        return output
    
    async def export_relationship_stats(self, user_id: int, db) -> io.BytesIO:
        """관계별 분석만 엑셀로 내보내기"""
        self.workbook = Workbook()
        
        # 기본 시트 제거
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])
        
        await self._create_relationship_breakdown_sheet(user_id, db)
        
        # 메모리 버퍼에 저장
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        
        return output
    
    async def export_personal_stats(self, user_id: int, db) -> io.BytesIO:
        """개인별 상세만 엑셀로 내보내기"""
        self.workbook = Workbook()
        
        # 기본 시트 제거
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])
        
        await self._create_personal_details_sheet(user_id, db)
        
        # 메모리 버퍼에 저장
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        
        return output
    
    async def export_events_stats(self, user_id: int, db) -> io.BytesIO:
        """이벤트별 기록만 엑셀로 내보내기"""
        self.workbook = Workbook()
        
        # 기본 시트 제거
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])
        
        await self._create_events_stats_sheet(user_id, db)
        
        # 메모리 버퍼에 저장
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        
        return output


# 서비스 인스턴스 생성
excel_export_service = ExcelExportService()